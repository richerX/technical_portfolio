''' 
openpyxl 
'''
import openpyxl
import datetime    
wb = openpyxl.load_workbook(filename = "abc.xlsx")  # open existing table
wb = openpyxl.Workbook()                            # create new table
sheet = wb.active                                   # grab the sheet
article = sheet['A1'].value                         # grab the value
sheet["A1"] = 42                                    # insert the value
wb.save('Olimp results.xlsx')                       # save table      
ws['A2'] = datetime.datetime.now()                  # all types will be automatically converted



''' 
xlwings (best for photos)
'''
import xlwings
filepath = "abc.xlsx"
wb = xlwings.Book(filepath)
sheet = wb.active
article = sheet.range("A1").value
sheet.range("A1").value = "article"



''' 
Google spreadsheets
'''

import gspread
from oauth2client.service_account import ServiceAccountCredentials

        
scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']  # do not touch
credentials = ServiceAccountCredentials.from_json_keyfile_name('DiscreteMathResults-c9f026c9cefb.json', scope)  # json file of account (https://github.com/burnash/gspread/blob/master/README.md)
gc = gspread.authorize(credentials)
url = 'https://docs.google.com/spreadsheets/d/1nac83TAP2AU4bjLvzF_wtGG5HbFgTx3QL5a9XTfUi4I/edit#gid=17124170'  # url to the spreadsheet
wb = gc.open_by_url(url)
sheets = wb.worksheets()
for sheet_open in sheets:
    sheet = wb.worksheet(sheet_open.title)
    matrix = sheet.get_all_values()  # we get all values on the list
